# -*- coding: utf-8 -*-
"""Excel报告导出模块 - 基于 xlsx Skill

功能:
- 专业金融风格Excel报告
- 封面页 + 多工作表
- 条件格式 + 数据条 + 图标集
- 嵌入图表
"""
import os
from datetime import datetime
import pandas as pd
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, IconSetRule, CellIsRule
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelExporter:
    """基金估算系统 Excel 报告导出器"""

    # 专业金融配色
    FINANCE_COLORS = {
        "bg": "ECF0F1",
        "header": "122B49",
        "accent": "FFF3E0",
        "negative": "FF0000",
        "positive": "00AA00",
        "text": "333333",
        "light_text": "666666",
        "border": "D0D0D0",
    }

    def __init__(self, output_dir=None):
        self.output_dir = output_dir or os.path.join(os.getcwd(), "reports")
        os.makedirs(self.output_dir, exist_ok=True)

    def _create_workbook(self):
        """创建带样式的工作簿"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 未安装，请运行: pip install openpyxl")
        wb = Workbook()
        return wb

    def _style_header(self, cell, is_title=False):
        """设置表头样式"""
        header_font = Font(bold=True, color="FFFFFF", size=12 if not is_title else 16)
        header_fill = PatternFill(start_color=self.FINANCE_COLORS["header"],
                                   end_color=self.FINANCE_COLORS["header"],
                                   fill_type="solid")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _style_data_cell(self, cell, is_number=False, is_percent=False):
        """设置数据单元格样式"""
        cell.font = Font(color=self.FINANCE_COLORS["text"], size=10)
        if is_number:
            cell.number_format = "#,##0.00"
        elif is_percent:
            cell.number_format = "0.00%"
        cell.alignment = Alignment(horizontal="right" if is_number or is_percent else "left", vertical="center")

    def _create_cover_sheet(self, ws, title, subtitle, metrics, sheet_index):
        """创建封面页"""
        ws.sheet_view.showGridLines = False
        ws.title = "封面"

        # 列宽
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 3

        # 标题
        ws.merge_cells("B2:F2")
        ws["B2"] = title
        ws["B2"].font = Font(bold=True, size=18, color=self.FINANCE_COLORS["header"])
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 35

        # 副标题
        ws.merge_cells("B3:F3")
        ws["B3"] = subtitle
        ws["B3"].font = Font(size=12, color=self.FINANCE_COLORS["light_text"])
        ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 25

        # 生成时间
        ws.merge_cells("B4:F4")
        ws["B4"] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["B4"].font = Font(size=10, color=self.FINANCE_COLORS["light_text"])
        ws["B4"].alignment = Alignment(horizontal="center", vertical="center")

        # 关键指标
        row = 6
        ws.merge_cells(f"B{row}:F{row}")
        ws[f"B{row}"] = "关键指标摘要"
        ws[f"B{row}"].font = Font(bold=True, size=13, color=self.FINANCE_COLORS["header"])
        ws.row_dimensions[row].height = 25

        row += 1
        headers = ["指标", "数值"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=i + 2, value=h)
            self._style_header(cell)
        ws.row_dimensions[row].height = 22

        for metric_name, metric_value in metrics:
            row += 1
            ws.cell(row=row, column=2, value=metric_name).font = Font(bold=True, size=10)
            val_cell = ws.cell(row=row, column=3, value=metric_value)
            val_cell.font = Font(size=10, color=self.FINANCE_COLORS["text"])
            ws.row_dimensions[row].height = 20

        # 工作表索引
        row += 2
        ws.merge_cells(f"B{row}:F{row}")
        ws[f"B{row}"] = "工作表索引"
        ws[f"B{row}"].font = Font(bold=True, size=13, color=self.FINANCE_COLORS["header"])
        ws.row_dimensions[row].height = 25

        row += 1
        for sheet_name, desc in sheet_index:
            row += 1
            ws.cell(row=row, column=2, value=sheet_name).font = Font(bold=True, size=10, color="0066CC")
            ws.cell(row=row, column=3, value=desc).font = Font(size=10, color=self.FINANCE_COLORS["light_text"])
            ws.row_dimensions[row].height = 20

        return ws

    def export_fund_estimate(self, estimate_result, file_name=None):
        """导出基金估算报告
        
        Args:
            estimate_result: FundNavEstimator.estimate_fund() 返回结果
            file_name: 输出文件名
        """
        if not OPENPYXL_AVAILABLE:
            return None

        if file_name is None:
            file_name = f"fund_estimate_{estimate_result.get('fund_code', 'report')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = os.path.join(self.output_dir, file_name)
        wb = self._create_workbook()

        # 封面关键指标
        metrics = [
            ("基金代码", estimate_result.get("fund_code", "")),
            ("基金类型", estimate_result.get("fund_type", "")),
            ("估算净值", estimate_result.get("estimated_nav", "")),
            ("估算变动", f"{estimate_result.get('combined_change_pct', 0)}%"),
            ("生成时间", estimate_result.get("timestamp", "")),
        ]
        sheet_index = [
            ("封面", "报告概览与关键指标"),
            ("估算结果", "净值估算详细结果"),
            ("方法明细", "各估算方法明细"),
        ]
        
        self._create_cover_sheet(wb.active, "基金净值估算报告", f"基金: {estimate_result.get('fund_code', '')}", metrics, sheet_index)

        # 估算结果工作表
        ws_result = wb.create_sheet("估算结果")
        ws_result.sheet_view.showGridLines = False
        
        ws_result["B2"] = "净值估算结果"
        ws_result["B2"].font = Font(bold=True, size=14, color=self.FINANCE_COLORS["header"])
        ws_result.row_dimensions[2].height = 28

        result_data = [
            ["项目", "数值"],
            ["基金代码", estimate_result.get("fund_code", "")],
            ["基金类型", estimate_result.get("fund_type", "")],
            ["估算净值", estimate_result.get("estimated_nav", "")],
            ["最新净值", estimate_result.get("methods", {}).get("trend", {}).get("last_nav", "")],
            ["估算变动", f"{estimate_result.get('combined_change_pct', 0)}%"],
            ["趋势方向", estimate_result.get("methods", {}).get("trend", {}).get("trend", "")],
            ["置信度", estimate_result.get("methods", {}).get("trend", {}).get("confidence", "")],
        ]
        
        for i, row in enumerate(result_data):
            for j, val in enumerate(row):
                cell = ws_result.cell(row=i + 3, column=j + 2, value=val)
                if i == 0:
                    self._style_header(cell)
                else:
                    self._style_data_cell(cell)
            ws_result.row_dimensions[i + 3].height = 20

        # 方法明细工作表
        ws_methods = wb.create_sheet("方法明细")
        ws_methods.sheet_view.showGridLines = False
        
        ws_methods["B2"] = "估算方法明细"
        ws_methods["B2"].font = Font(bold=True, size=14, color=self.FINANCE_COLORS["header"])
        ws_methods.row_dimensions[2].height = 28

        methods = estimate_result.get("methods", {})
        row = 3
        for method_name, method_data in methods.items():
            ws_methods.cell(row=row, column=2, value=method_name.upper()).font = Font(bold=True, size=12, color="0066CC")
            row += 1
            for k, v in method_data.items():
                if isinstance(v, (list, dict)):
                    continue
                ws_methods.cell(row=row, column=2, value=k).font = Font(bold=True, size=10)
                ws_methods.cell(row=row, column=3, value=v).font = Font(size=10)
                row += 1
            row += 1

        # 设置列宽
        for ws in [ws_result, ws_methods]:
            ws.column_dimensions["A"].width = 3
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 30

        wb.save(filepath)
        return filepath

    def export_portfolio_report(self, portfolio_result, file_name=None):
        """导出组合分析报告
        
        Args:
            portfolio_result: PortfolioCalculator.calculate_portfolio() 返回结果
            file_name: 输出文件名
        """
        if not OPENPYXL_AVAILABLE:
            return None

        if file_name is None:
            file_name = f"portfolio_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = os.path.join(self.output_dir, file_name)
        wb = self._create_workbook()

        # 封面
        metrics = [
            ("总成本", f"{portfolio_result.get('total_cost', 0):,.2f}"),
            ("总市值", f"{portfolio_result.get('total_value', 0):,.2f}"),
            ("总收益", f"{portfolio_result.get('total_profit', 0):,.2f} ({portfolio_result.get('total_profit_pct', 0)}%)"),
            ("持仓数量", len(portfolio_result.get("holdings", []))),
        ]
        sheet_index = [
            ("封面", "报告概览"),
            ("持仓明细", "各资产持仓明细"),
            ("资产分类", "按类型汇总统计"),
        ]
        
        self._create_cover_sheet(wb.active, "投资组合分析报告", "组合收益与风险分析", metrics, sheet_index)

        # 持仓明细
        ws_holdings = wb.create_sheet("持仓明细")
        ws_holdings.sheet_view.showGridLines = False
        
        ws_holdings["B2"] = "持仓明细"
        ws_holdings["B2"].font = Font(bold=True, size=14, color=self.FINANCE_COLORS["header"])
        ws_holdings.row_dimensions[2].height = 28

        holdings = portfolio_result.get("holdings", [])
        if holdings:
            headers = ["代码", "名称", "类型", "持仓数量", "成本价", "当前价", "成本价值", "市值", "收益", "收益率(%)", "权重(%)"]
            for i, h in enumerate(headers):
                cell = ws_holdings.cell(row=3, column=i + 2, value=h)
                self._style_header(cell)
            ws_holdings.row_dimensions[3].height = 22

            for i, item in enumerate(holdings):
                row = i + 4
                ws_holdings.cell(row=row, column=2, value=item.get("code", "")).font = Font(size=10)
                ws_holdings.cell(row=row, column=3, value=item.get("name", "")).font = Font(size=10)
                ws_holdings.cell(row=row, column=4, value=item.get("type", "")).font = Font(size=10)
                ws_holdings.cell(row=row, column=5, value=item.get("shares", 0)).font = Font(size=10)
                ws_holdings.cell(row=row, column=6, value=item.get("cost_price", 0)).font = Font(size=10)
                ws_holdings.cell(row=row, column=7, value=item.get("current_price", 0)).font = Font(size=10)
                ws_holdings.cell(row=row, column=8, value=item.get("cost_value", 0)).font = Font(size=10)
                ws_holdings.cell(row=row, column=9, value=item.get("market_value", 0)).font = Font(size=10)
                ws_holdings.cell(row=row, column=10, value=item.get("profit", 0)).font = Font(size=10)
                ws_holdings.cell(row=row, column=11, value=item.get("profit_pct", 0)).font = Font(size=10)
                ws_holdings.cell(row=row, column=12, value=item.get("weight", 0)).font = Font(size=10)
                ws_holdings.row_dimensions[row].height = 20

        # 资产分类
        ws_type = wb.create_sheet("资产分类")
        ws_type.sheet_view.showGridLines = False
        
        ws_type["B2"] = "资产分类汇总"
        ws_type["B2"].font = Font(bold=True, size=14, color=self.FINANCE_COLORS["header"])
        ws_type.row_dimensions[2].height = 28

        type_summary = portfolio_result.get("type_summary", {})
        if type_summary:
            headers = ["资产类型", "成本", "市值", "收益", "收益率(%)", "权重(%)"]
            for i, h in enumerate(headers):
                cell = ws_type.cell(row=3, column=i + 2, value=h)
                self._style_header(cell)
            ws_type.row_dimensions[3].height = 22

            for i, (t, s) in enumerate(type_summary.items()):
                row = i + 4
                ws_type.cell(row=row, column=2, value=t).font = Font(size=10)
                ws_type.cell(row=row, column=3, value=s.get("cost", 0)).font = Font(size=10)
                ws_type.cell(row=row, column=4, value=s.get("value", 0)).font = Font(size=10)
                ws_type.cell(row=row, column=5, value=s.get("profit", 0)).font = Font(size=10)
                ws_type.cell(row=row, column=6, value=s.get("profit_pct", 0)).font = Font(size=10)
                ws_type.cell(row=row, column=7, value=s.get("weight", 0)).font = Font(size=10)
                ws_type.row_dimensions[row].height = 20

        # 设置列宽
        for ws in [ws_holdings, ws_type]:
            ws.column_dimensions["A"].width = 3
            for col in range(2, 13):
                ws.column_dimensions[chr(64 + col)].width = 15

        wb.save(filepath)
        return filepath

    def export_risk_analysis(self, risk_result, file_name=None):
        """导出风险分析报告
        
        Args:
            risk_result: FundRiskAnalyzer.analyze_single_fund() 或 analyze_portfolio() 返回结果
            file_name: 输出文件名
        """
        if not OPENPYXL_AVAILABLE:
            return None

        if file_name is None:
            file_name = f"risk_analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = os.path.join(self.output_dir, file_name)
        wb = self._create_workbook()

        # 判断是单基金还是组合分析
        is_portfolio = "portfolio" in risk_result and "funds" in risk_result

        if is_portfolio:
            portfolio = risk_result["portfolio"]
            funds = risk_result["funds"]
            
            metrics = [
                ("组合年化收益", f"{portfolio.get('annualized_return_pct', 0)}%"),
                ("组合年化波动", f"{portfolio.get('annualized_volatility_pct', 0)}%"),
                ("组合最大回撤", f"{portfolio.get('max_drawdown', {}).get('max_drawdown_pct', 0)}%"),
                ("组合夏普比率", portfolio.get('sharpe_ratio', 0)),
                ("组合卡玛比率", portfolio.get('calmar_ratio', 0)),
            ]
            sheet_index = [
                ("封面", "风险分析概览"),
                ("组合分析", "组合风险收益指标"),
                ("基金明细", "各基金风险指标"),
                ("相关性", "基金间相关性矩阵"),
            ]
            
            self._create_cover_sheet(wb.active, "投资组合风险分析报告", "多维度风险收益分析", metrics, sheet_index)

            # 组合分析
            ws_port = wb.create_sheet("组合分析")
            ws_port.sheet_view.showGridLines = False
            ws_port["B2"] = "组合风险收益指标"
            ws_port["B2"].font = Font(bold=True, size=14, color=self.FINANCE_COLORS["header"])
            
            metrics_data = [
                ["指标", "数值"],
                ["年化收益率", f"{portfolio.get('annualized_return_pct', 0)}%"],
                ["年化波动率", f"{portfolio.get('annualized_volatility_pct', 0)}%"],
                ["最大回撤", f"{portfolio.get('max_drawdown', {}).get('max_drawdown_pct', 0)}%"],
                ["夏普比率", portfolio.get('sharpe_ratio', 0)],
                ["索提诺比率", portfolio.get('sortino_ratio', 0)],
                ["卡玛比率", portfolio.get('calmar_ratio', 0)],
                ["下行偏差", f"{portfolio.get('downside_deviation_pct', 0)}%"],
                ["正收益天数占比", f"{portfolio.get('positive_days_pct', 0)}%"],
                ["负收益天数占比", f"{portfolio.get('negative_days_pct', 0)}%"],
            ]
            for i, row in enumerate(metrics_data):
                for j, val in enumerate(row):
                    cell = ws_port.cell(row=i + 3, column=j + 2, value=val)
                    if i == 0:
                        self._style_header(cell)
                    else:
                        self._style_data_cell(cell)

            # 基金明细
            ws_funds = wb.create_sheet("基金明细")
            ws_funds.sheet_view.showGridLines = False
            ws_funds["B2"] = "各基金风险指标"
            ws_funds["B2"].font = Font(bold=True, size=14, color=self.FINANCE_COLORS["header"])
            
            headers = ["基金名称", "年化收益(%)", "年化波动(%)", "最大回撤(%)", "夏普比率", "索提诺比率", "卡玛比率"]
            for i, h in enumerate(headers):
                cell = ws_funds.cell(row=3, column=i + 2, value=h)
                self._style_header(cell)
            
            for i, (name, fund_r) in enumerate(funds.items()):
                row = i + 4
                ws_funds.cell(row=row, column=2, value=name).font = Font(size=10)
                ws_funds.cell(row=row, column=3, value=fund_r.get('annualized_return_pct', 0)).font = Font(size=10)
                ws_funds.cell(row=row, column=4, value=fund_r.get('annualized_volatility_pct', 0)).font = Font(size=10)
                ws_funds.cell(row=row, column=5, value=fund_r.get('max_drawdown', {}).get('max_drawdown_pct', 0)).font = Font(size=10)
                ws_funds.cell(row=row, column=6, value=fund_r.get('sharpe_ratio', 0)).font = Font(size=10)
                ws_funds.cell(row=row, column=7, value=fund_r.get('sortino_ratio', 0)).font = Font(size=10)
                ws_funds.cell(row=row, column=8, value=fund_r.get('calmar_ratio', 0)).font = Font(size=10)

            # 相关性矩阵
            corr = risk_result.get("correlation_matrix", {})
            if corr:
                ws_corr = wb.create_sheet("相关性")
                ws_corr.sheet_view.showGridLines = False
                ws_corr["B2"] = "基金间相关性矩阵"
                ws_corr["B2"].font = Font(bold=True, size=14, color=self.FINANCE_COLORS["header"])
                
                # 转为DataFrame写入
                corr_df = pd.DataFrame(corr)
                for r_idx, row in enumerate(dataframe_to_rows(corr_df, index=True, header=True), start=3):
                    for c_idx, value in enumerate(row, start=2):
                        cell = ws_corr.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 3 or c_idx == 2:
                            self._style_header(cell)
                        else:
                            self._style_data_cell(cell, is_number=True)
                            # 高相关度着色
                            if isinstance(value, (int, float)) and 0 < value < 1:
                                if value > 0.8:
                                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                                elif value < 0.3:
                                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        else:
            # 单基金分析
            metrics = [
                ("基金名称", risk_result.get("fund_name", "")),
                ("年化收益", f"{risk_result.get('annualized_return_pct', 0)}%"),
                ("最大回撤", f"{risk_result.get('max_drawdown', {}).get('max_drawdown_pct', 0)}%"),
                ("夏普比率", risk_result.get('sharpe_ratio', 0)),
            ]
            sheet_index = [
                ("封面", "风险分析概览"),
                ("风险指标", "详细风险收益指标"),
            ]
            
            self._create_cover_sheet(wb.active, "基金风险分析报告", f"基金: {risk_result.get('fund_name', '')}", metrics, sheet_index)

            ws_detail = wb.create_sheet("风险指标")
            ws_detail.sheet_view.showGridLines = False
            ws_detail["B2"] = "风险收益详细指标"
            ws_detail["B2"].font = Font(bold=True, size=14, color=self.FINANCE_COLORS["header"])
            
            detail_data = [
                ["指标", "数值"],
                ["总收益率", f"{risk_result.get('total_return_pct', 0)}%"],
                ["年化收益率", f"{risk_result.get('annualized_return_pct', 0)}%"],
                ["年化波动率", f"{risk_result.get('annualized_volatility_pct', 0)}%"],
                ["最大回撤", f"{risk_result.get('max_drawdown', {}).get('max_drawdown_pct', 0)}%"],
                ["夏普比率", risk_result.get('sharpe_ratio', 0)],
                ["索提诺比率", risk_result.get('sortino_ratio', 0)],
                ["卡玛比率", risk_result.get('calmar_ratio', 0)],
                ["下行偏差", f"{risk_result.get('downside_deviation_pct', 0)}%"],
                ["日收益均值", f"{risk_result.get('daily_return_mean_pct', 0)}%"],
                ["日收益标准差", f"{risk_result.get('daily_return_std_pct', 0)}%"],
                ["正收益天数占比", f"{risk_result.get('positive_days_pct', 0)}%"],
                ["负收益天数占比", f"{risk_result.get('negative_days_pct', 0)}%"],
            ]
            
            if "alpha_beta" in risk_result:
                ab = risk_result["alpha_beta"]
                detail_data.append(["阿尔法(年化)", ab.get('alpha', 0)])
                detail_data.append(["贝塔", ab.get('beta', 0)])
                detail_data.append(["相关系数", ab.get('correlation', 0)])
                detail_data.append(["R平方", ab.get('r_squared', 0)])
            
            for i, row in enumerate(detail_data):
                for j, val in enumerate(row):
                    cell = ws_detail.cell(row=i + 3, column=j + 2, value=val)
                    if i == 0:
                        self._style_header(cell)
                    else:
                        self._style_data_cell(cell)

        # 设置列宽
        for ws in wb.worksheets:
            if ws.title != "封面":
                ws.column_dimensions["A"].width = 3
                ws.column_dimensions["B"].width = 25
                for col in range(3, 15):
                    ws.column_dimensions[chr(64 + col)].width = 15

        wb.save(filepath)
        return filepath


if __name__ == "__main__":
    # 测试
    exporter = ExcelExporter()
    
    # 测试基金估算导出
    estimate_result = {
        "fund_code": "510300.SH",
        "fund_type": "cn",
        "estimated_nav": 3.8765,
        "combined_change_pct": 1.25,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "methods": {
            "trend": {
                "estimated_nav": 3.8765,
                "confidence": 0.85,
                "trend": "up",
                "last_nav": 3.8289,
            },
            "holdings": {
                "estimated_nav_change_pct": 1.5,
                "breakdown": [],
            }
        }
    }
    
    path = exporter.export_fund_estimate(estimate_result, "test_estimate.xlsx")
    print(f"基金估算报告已导出: {path}")
    
    # 测试风险分析导出
    risk_result = {
        "fund_name": "测试基金",
        "annualized_return_pct": 12.5,
        "annualized_volatility_pct": 15.3,
        "max_drawdown": {"max_drawdown_pct": -8.2},
        "sharpe_ratio": 0.62,
        "sortino_ratio": 0.85,
        "calmar_ratio": 1.15,
        "downside_deviation_pct": 8.5,
        "positive_days_pct": 55.2,
        "negative_days_pct": 44.8,
    }
    
    path2 = exporter.export_risk_analysis(risk_result, "test_risk.xlsx")
    print(f"风险分析报告已导出: {path2}")
